import pandas as pd
import os
import xlsxwriter
import openpyxl
from datetime import datetime

def create_output_template(workbook, worksheet):
    """יוצר את תבנית קובץ הפלט"""
    # הגדרת פורמטים
    header_format = workbook.add_format({
        'bold': True,
        'align': 'center',
        'bg_color': '#D9D9D9'
    })
    
    green_format = workbook.add_format({
        'bg_color': '#C6E0B4'
    })
    
    blue_format = workbook.add_format({
        'bg_color': '#BDD7EE'
    })
    
    yellow_format = workbook.add_format({
        'bg_color': '#FFE699'
    })

    # Row 1 - חותרות חברות
    worksheet.write(0, 0, 'חברה', header_format)
    worksheet.write(0, 1, 'רומי')
    worksheet.write(0, 2, 'רומי') 
    worksheet.write(0, 3, 'החזקות')
    worksheet.write(0, 4, 'החזקות')
    worksheet.write(0, 5, 'נדלן')
    worksheet.write(0, 6, 'נדלן')
    worksheet.write(0, 7, 'נדלן')
    worksheet.write(0, 8, 'יובל פרטי')
    worksheet.write(0, 9, 'יובל פרטי')
    
    # Row 2 - מסגרות
    worksheet.write(1, 0, 'מסגרות', yellow_format)
    worksheet.write(1, 1, '-')
    worksheet.write(1, 2, '-')
    worksheet.write(1, 3, 20000)
    worksheet.write(1, 4, '-')
    worksheet.write(1, 5, 100000)
    worksheet.write(1, 6, 20000)
    worksheet.write(1, 7, 50000)
    worksheet.write(1, 8, 70500)
    worksheet.write(1, 9, 0)
    # Calculate total of numeric values, replacing '-' with 0
    total = sum(0 if x == '-' else x for x in [0, 20000, 0, 100000, 20000, 50000, 70500, 0])
    worksheet.write(1, 10, total)
    # Row 3 - שם הבנק
    worksheet.write(2, 0, 'שם הבנק', header_format)
    worksheet.write(2, 1, 'מזרחי')
    worksheet.write(2, 2, 'מזרחי')
    worksheet.write(2, 3, 'פועלים')
    worksheet.write(2, 4, 'מזרחי')
    worksheet.write(2, 5, 'מזרחי')
    worksheet.write(2, 6, 'דיסקונט')
    worksheet.write(2, 7, 'פועלים')
    worksheet.write(2, 8, 'פועלים')
    worksheet.write(2, 9, 'פועלים')
    
    # Row 4 - מספר ח-ן
    account_numbers = ['193744', '197154', '31324', '177315', '172615', '153771129', '313222', '409937', '55533']
    worksheet.write(3, 0, 'מספר ח-ן', header_format)
    for col, number in enumerate(account_numbers, start=1):
        worksheet.write(3, col, number, green_format)
    
    worksheet.right_to_left()
    return account_numbers

def find_account_number(sheet):
    """מוצא את מספר החשבון בגיליון"""
    for row in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row=row, column=2).value
        if cell_value and 'מזרחי' in str(cell_value):
            account_number = str(cell_value).split()[-1]
            print(f"נמצא מספר חשבון: {account_number} בשורה {row} עמודה 2")
            return account_number
    raise ValueError("לא נמצא מספר חשבון בקובץ")

def find_output_column(account_number, account_numbers):
    """מוצא את העמודה המתאימה בקובץ הפלט"""
    for col, acc in enumerate(account_numbers, start=1):
        if acc == account_number:
            print(f"נמצאה התאמה בעמודה {col} בקובץ הפלט")
            return col
    raise ValueError(f"לא נמצא חשבון מתאים {account_number}")

def process_bank_file(input_file, worksheet, account_numbers, workbook):
    input_wb = openpyxl.load_workbook(input_file)
    input_sheet = input_wb.active
    
    # מציאת מספר חשבון והעמודה המתאימה
    account_number = find_account_number(input_sheet)
    output_col = find_output_column(account_number, account_numbers)
    
    # מילון לשמירת היתרה האחרונה לכל תאריך
    balances_by_date = {}
    
    # קריאת כל היתרות ושמירת האחרונה לכל תאריך
    for row in range(10, input_sheet.max_row + 1):
        date_cell = input_sheet.cell(row=row, column=1).value
        balance_cell = input_sheet.cell(row=row, column=10).value
        
        if isinstance(date_cell, str) and isinstance(balance_cell, (int, float)):
            balances_by_date[date_cell] = balance_cell
    
    # המרת המפתחות למילון חדש עם תאריכים במקום מחרוזות
    sorted_dates = {}
    for date_str, balance in balances_by_date.items():
        try:
            actual_date = datetime.strptime(date_str, '%d/%m/%y')
            sorted_dates[actual_date] = balance
        except ValueError as e:
            print(f"שגיאה בהמרת תאריך {date_str}: {e}")
            continue
    
    # כתיבת היתרות לקובץ הפלט
    next_row = 4
    date_format = workbook.add_format({'num_format': 'dd/mm/yy'})
    number_format = workbook.add_format({'num_format': '#,##0.00'})
    
    # מיון לפי תאריכים אמיתיים
    for date in sorted(sorted_dates.keys()):
        balance = sorted_dates[date]
        print(f"כותב לשורה {next_row + 1}:")
        print(f"תאריך {date.strftime('%d/%m/%y')} בעמודה A")
        print(f"יתרה {balance} בעמודה {output_col + 1}")
        
        try:
            worksheet.write_datetime(next_row, 0, date, date_format)
            worksheet.write_number(next_row, output_col, balance, number_format)
            next_row += 1
        except Exception as e:
            print(f"שגיאה בכתיבה: {str(e)}")
            raise
    
    input_wb.close()
    return True

def find_bank_files():
    """מוצא את כל קבצי הבנק בתיקיית bank_statements"""
    bank_dir = 'bank_statements'
    print(f"מחפש קבצים בתיקייה: {os.path.abspath(bank_dir)}")
    
    if not os.path.exists(bank_dir):
        print("✗ תיקיית bank_statements לא קיימת")
        return []
    
    files = [f for f in os.listdir(bank_dir) if f.endswith('.xlsx')]
    print(f"נמצאו {len(files)} קבצים: {files}")
    
    return [os.path.join(bank_dir, f) for f in files]

def process_bank_files():
    # יצירת תיקיית output אם לא קיימת
    if not os.path.exists('output'):
        os.makedirs('output')
        
    workbook = xlsxwriter.Workbook('output/output.xlsx')
    worksheet = workbook.add_worksheet()
    
    try:
        account_numbers = create_output_template(workbook, worksheet)
        
        for input_file in find_bank_files():
            try:
                print(f"\nמעבד קובץ: {os.path.basename(input_file)}")
                process_bank_file(input_file, worksheet, account_numbers, workbook)
            except Exception as e:
                print(f"✗ שגיאה בקובץ {os.path.basename(input_file)}:")
                print(f"  פירוט: {str(e)}")
    finally:
        workbook.close()

def main():
    process_bank_files()
    print("\nהתהליך הסתיים")

if __name__ == "__main__":
    main()